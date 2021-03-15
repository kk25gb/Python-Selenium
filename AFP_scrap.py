from selenium import webdriver
import openpyxl


wb =openpyxl.Workbook()
# wb = openpyxl.load_workbook("AFP_record.xlsx")
ws = wb["Sheet"]
title_list = []
ws['A1'] = '來源處'
ws['B1'] = '主題'
ws['C1'] = '日期'
ws['D1'] = '檔名'
# for cell in ws['B']:
#     title_list.append(cell.value[:41])

# use webdriver to open chrome open google search page
driver = webdriver.Chrome()


driver.get('https://www.afp.com/en')
driver.implicitly_wait(20)
main_values = ['COVID-19 ', 'coronavirus disease ', 'epidemic prevention ', 'pandemic ', 'society ']
pending_values = ['citizen-state relations ', 'civil society ', 'civic participation/civic engagement ', 'trust/trustworthiness ',
          'confidence ', 'reciprocity ', 'cooperation/collaboration ', 'cooperate/collaborate ', 'public affairs ']


row_index = 2
txtFile_count = 1
for j in range(len(main_values)):
    for i in range(len(pending_values)):
        driver.implicitly_wait(15)
        search_btn = driver.find_element_by_id("buttonsearch")
        search_btn.click()

        driver.implicitly_wait(10)
        text_area = driver.find_element_by_name("search_block_form")
        driver.implicitly_wait(10)
        # 定義搜尋關鍵字
        searchKey = "Taiwan "+main_values[j]+pending_values[i]
        print(searchKey)
        text_area.send_keys(searchKey)
        driver.find_element_by_id('edit-submit').click()

        driver.implicitly_wait(10)
        driver.find_element_by_link_text("Date").click()


        last_date = ""
        valid_date = True
        while valid_date:
            for k in range(10):
                driver.implicitly_wait(15)
                # 抓日期
                date = driver.find_element_by_xpath('//*[@id="contentcol"]/div[3]/div['+str(k+1)+']/div/span/strong').text
                # 存最後一個抓取的文章之日期
                last_date = date[:10]
                # 判斷日期合法性
                if (date[6:10] == "2020") and ((int(date[:2]) == 1 and int(date[3:5]) >= 25) or (int(date[:2]) > 1)):
                    title = driver.find_element_by_css_selector\
                    ("#contentcol > div:nth-child(3) > div:nth-child("+str(k+1)+") > h4 > span > span > a").text
                    # 抓取標題下的簡介
                    introStr = driver.find_element_by_css_selector(
                        '#contentcol > div:nth-child(3) > div:nth-child(' + str(k + 1) + ') > p').text
                    valid_content = False
                    # 檢查文章內是否有搜到Taiwan
                    if "Taiwan" in introStr:
                        valid_content = True
                    # 過濾不要的title 與沒有Taiwan的文章
                    if ("videos" not in title) and ("Contact" not in title) and ("AFP trainee scheme" not in title) \
                            and ("Newswire" not in title) and ("Business Wire" != title) and (title != '') and (title != 'Documentation for topics ') \
                            and (title != 'Medialab') and (title[:41] not in title_list) and valid_content:
                        # print("write into excel" +'txt:'+str(txtFile_count)+"row:"+str(row_index))
                        ws['A' + str(row_index)] = "法新社"
                        ws['B' + str(row_index)] = title
                        ws['C' + str(row_index)] = date[:10]
                        title_list.append(title[:41])
                        # click into article
                        driver.find_element_by_css_selector\
                        ("#contentcol > div:nth-child(3) > div:nth-child("+str(k+1)+") > h4 > span > span > a").click()
                        # get the content
                        content = driver.find_element_by_css_selector\
                        ("#contentcol > div > div.article_content.line.linemam.mb2 > div.w85.right.txt12.txtlh18.txtblack.txtjustify.textcontent").text
                        if content == '':
                            # print("clean last"+'txt:'+str(txtFile_count)+"row:"+str(row_index))
                            ws['A' + str(row_index)] = ''
                            ws['B' + str(row_index)] = ''
                            ws['C' + str(row_index)] = ''
                        else:
                            file_name = 'AFP' + str(txtFile_count)
                            f = open('Records/' + file_name + '.txt', 'w', encoding="utf-8")
                            f.write(content)
                            ws['D' + str(row_index)] = file_name
                            print(date[:10] +' '+ title[:41] +' '+ file_name +'txt:'+str(txtFile_count)+"row:"+str(row_index))
                            txtFile_count += 1
                            row_index += 1
                        f.close()
                        driver.back()
                        wb.save("record.xlsx")
                else:
                    valid_date = False
                    print("tested invalid")
                    break
            if (last_date[6:10] == "2020") and ((int(last_date[:2]) == 1 and int(last_date[3:5]) >= 25) or (int(last_date[:2]) > 1)):
                driver.find_element_by_css_selector("#contentcol > div.pagination.line.mb2.linemam > ul > li.pager-next.btn.txtwhite.bgblue > a > span").click()
                print("next page")


driver.close()