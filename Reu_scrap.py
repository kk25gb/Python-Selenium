from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
# from selenium.webdriver.common.keys import Keys
# from selenium.webdriver.support import expected_conditions as EC
# from selenium.webdriver.common.by import By
import time
import random
import openpyxl


# 開啟record excel檔
wb = openpyxl.load_workbook("REU_NewRecords.xlsx")
ws = wb["Sheet1"]
# 創建存放標題的串列
title_list = []
# 設定excel的起始index和檔名index
txtFile_count = 1224
row_index = txtFile_count + 1

for cell in ws['B']:
    title_list.append(cell.value[:41].lower())

month = {'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5, 'JUN': 6, 'JUL': 7, 'AUG': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11,
     'DEC': 12}

# use webdriver to open chrome open google search page
driver = webdriver.Chrome()
driver.get('https://uk.reuters.com/search/news?blob=Taiwan+covid-19&sortBy=relevance&dateRange=all')
driver.maximize_window()
driver.implicitly_wait(10)

result_num = driver.find_element_by_xpath('//*[@id="content"]/section[2]/div/div[1]/div[4]/div/div[1]/span[1]').text
rnum = int(result_num)
print(rnum)

wasInPage = False
for k in range(rnum+1):
    if k == 3:
        continue

    time.sleep(random.randint(1, 3))
    print('pointer:', k+1)

    try:
        href = driver.find_element_by_xpath\
            ('//*[@id="content"]/section[2]/div/div[1]/div[4]/div/div[3]/div[' + str(k + 1) + ']/div/h3/a').get_attribute('href')
    except:
        load_btn = driver.find_element_by_xpath('//*[@id="content"]/section[2]/div/div[1]/div[4]/div/div[4]/div[1]')
        ActionChains(driver).move_to_element(load_btn).click().perform()
        print('load')

    # get date
    date = driver.find_element_by_css_selector('#content > section:nth-child(5) > div > div.column1.col.col-10 > div.module > div > div.search-result-list.news-search > div:nth-child(' + str(k + 1) + ') > div.search-result-content > h5').text
    # convert date format
    date = date.split(" ")
    date[0] = month[date[0][:3]]
    date[1] = int(date[1][:len(date[1]) - 1])
    date[2] = int(date[2])
    date_s = str(date[0]) + '/' + str(date[1]) + '/' + str(date[2])
    # check date
    if (date[2] == 2020) and ((date[0] == 1 and (date[1]) >= 25) or (date[0] > 1)):
        title = driver.find_element_by_xpath(
            '//*[@id="content"]/section[2]/div/div[1]/div[4]/div/div[3]/div['+str(k+1)+']/div/h3/a').text
        if ('REUTERS NEWS SCHEDULE' not in title) and ('EMERGING MARKETS' not in title) and \
        ('FACTBOX' not in title) and ('REFILE - EMERGING MARKETS' not in title) and (title[:41].lower() not in title_list):

            time.sleep(random.randint(1, 2))
            # 進入文章頁面
            href = driver.find_element_by_xpath\
                ('//*[@id="content"]/section[2]/div/div[1]/div[4]/div/div[3]/div[' + str(k + 1) + ']/div/h3/a').get_attribute('href')
            driver.execute_script('window.open("' + href + '");')
            driver.switch_to.window(driver.window_handles[1])

            content = driver.find_element_by_xpath('//*[@id="__next"]/div/div[4]/div[1]/article/div[1]').text

            # 過濾沒有Taiwan的文章
            if 'Taiwan' in content and 'covid-19' in content and content != '':
                title_list.append(title[0:41].lower())
                ws['A' + str(row_index)] = "路透社"
                ws['B' + str(row_index)] = title
                ws['C' + str(row_index)] = date_s

                file_name = 'REU_new_' + str(txtFile_count)
                f = open('REU_new/' + file_name + '.txt', 'w', encoding="utf-8")

                time.sleep(random.randint(1, 2))
                paragraph_list = driver.find_elements_by_xpath('//*[@id="__next"]/div/div[4]/div[1]/article/div[1]/p')
                for w in paragraph_list:
                    try:
                        if w.text[0] == '*':
                            continue
                        else:
                            f.write(w.text + '\n')
                    except:
                        f.write(w.text + '\n')
                ws['D' + str(row_index)] = file_name

                print(date_s + ' ' + title[:41] + ' ' + file_name)
                txtFile_count += 1
                row_index += 1
                f.close()
                wb.save("REU_NewRecords.xlsx")

                time.sleep(random.randint(1, 3))
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                # wasInPage = True
            else:
                print('invalid content')
                # wasInPage = True
                time.sleep(random.randint(1, 2))
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
        else:
            print('invalid title:' + title)
    else:
        print("invalid date:" + date_s)
        break

# driver.close()