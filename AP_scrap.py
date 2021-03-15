from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from time import sleep
import openpyxl
# 開啟record excel檔
wb = openpyxl.load_workbook("APrecord.xlsx")
ws = wb["Sheet1"]
# 創建存放標題的串列
title_list = []
for cell in ws['B']:
    title_list.append(cell.value[:41])

# 設定excel的起始index和檔名index
txtFile_count = 15
row_index = txtFile_count + 1

driver = webdriver.Chrome()
# 打開首頁
driver.get('https://apnews.com/')
driver.implicitly_wait(5)

sleep(20)
webdriver.ActionChains(driver).send_keys(Keys.ESCAPE).perform()

search_btn = driver.find_element_by_css_selector\
    ('#root > div > main > div.Header > div.fluid-wrapper > nav > ol > li.header-navigation-item.search > div > svg')

# 輸入搜尋關鍵字 送出
text_area = driver.find_element_by_xpath('//*[@id="root"]/div/main/div[1]/div[2]/div/div/input')
text_area.send_keys('Taiwan covid-19')

# 檢查總共多少筆資料
search_sec = driver.find_element_by_css_selector\
('#root > div > main > div.Header > div.search-wrapper > div > div > section > section:nth-child(2) > h2').text
print(search_sec)
if search_sec == 'STORIES':
    result = driver.find_elements_by_xpath\
        ('//*[@id="root"]/div/main/div[1]/div[2]/div/div/section/section[2]/div/div/ul/li')
else:
    result = []
print(len(result))

# 若需要expand 則點擊
if len(result) > 5:
    driver.find_element_by_xpath('//*[@id="root"]/div/main/div[1]/div[2]/div/div/section/section[2]/div/div[2]').click()


for i in range(len(result)):
    # 定需要找的資料為目標
    target = driver.find_element_by_css_selector \
    ('#root>div>main>div.Header>div.search-wrapper > div > div > section > section:nth-child(2) > div > div > ul > li:nth-child('+str(i+1)+')>a')

    # 檢查為外部連結 跳過不點擊
    if target.get_attribute('target') == '_blank':
        print('external link')
        continue

    href = driver.find_element_by_xpath('//*[@id="root"]/div/main/div[1]/div[2]/div/div/section/section[2]/div/div/ul/li['+str(i+1)+']/a').get_attribute('href')
    driver.execute_script('window.open("' + href + '");')
    driver.switch_to.window(driver.window_handles[1])
    # 抓取日期
    sleep(1)
    date_ori = driver.find_element_by_class_name('Timestamp').get_attribute("data-source")
    date = date_ori[:10].replace("-", "/")
    # 檢測日期
    if date[:4] == "2020" and ((int(date[5:7]) == 1 and int(date[8:10]) >= 25) or (int(date[5:7]) > 1)):
        # 取標題與內容
        content = driver.find_element_by_css_selector('#root > div > main > div.Body > div > div.Article').text
        title = driver.find_element_by_css_selector('#root > div > main > div.Body > div > div.CardHeadline > div > h1').text
        # 檢查是否與台灣有關 且非重複
        if (content != '') and ("Taiwan" in content) and ('covid-19' in content) and (title[:41] not in title_list):
            title_list.append(title[:41])
            ws['A' + str(row_index)] = "美聯社"
            ws['B' + str(row_index)] = title
            ws['C' + str(row_index)] = date

            file_name = 'AP' + str(txtFile_count)
            f = open('Records/' + file_name + '.txt', 'w', encoding="utf-8")
            f.write(content)
            ws['D' + str(row_index)] = file_name
            print(date + ' ' + title[:41] + ' ' + file_name + 'txt:' + str(txtFile_count) + "row:" + str(row_index))
            # 寫入檔案完成 各指標+1 關閉檔案
            txtFile_count += 1
            row_index += 1
            f.close()
        else:
            print("invalid content")
        # 儲存excel
        wb.save("APrecord.xlsx")
        driver.back()
        # 再點開搜尋結果之隱藏div 以便找下一筆搜尋結果
        search_btn.click()
        print('B search_btn clicked')
    else:
        print("invalid date")
        driver.back()
        # 再點開搜尋結果之隱藏div
        search_btn.click()
        print('C search_btn clicked')



# driver.close()